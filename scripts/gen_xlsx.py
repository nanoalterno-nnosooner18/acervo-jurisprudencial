# -*- coding: utf-8 -*-
"""Genera la versión Excel .xlsx del acervo con autofiltros y hojas por materia."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from acervo_dataset import ACERVO
from gen_html import construir_entradas_normalizadas

ACENTO = "6B1818"
BEIGE = "EFEAE0"
CREMA = "FAF8F4"
TINTA_SUAVE = "555555"

COLUMNAS = [
    ("id",                "ID",                12),
    ("jerarquia",         "Niv.",              5),
    ("instancia_corta",   "Instancia",         18),
    ("materia_principal", "Materia",           18),
    ("tema",              "Tema",              30),
    ("subtema",           "Subtema",           28),
    ("tipo",              "Tipo",              16),
    ("epoca",             "Época",             14),
    ("vigencia",          "Vigencia",          16),
    ("ambito",            "Ámbito",            22),
    ("rubro",             "Rubro",             60),
    ("registro_digital",  "Registro SJF",      24),
    ("tipo_resolucion",   "Resolución",        24),
    ("fecha_publicacion", "Fecha",             10),
    ("fuente_publicacion","Fuente",            28),
    ("votacion",          "Votación",          18),
    ("texto_resumen",     "Texto",             70),
    ("trascendencia",     "Trascendencia",     50),
    ("explicacion",       "Explicación",       60),
    ("tesis_clave",       "Tesis (clave)",     22),
    ("url_sjf",           "URL SJF / SCJN",    40),
    ("etiquetas",         "Etiquetas",         32),
]


def estilo_encabezado(cell):
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=ACENTO)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        bottom=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
    )


def estilo_celda(cell, alt=False):
    cell.font = Font(name="Calibri", size=10)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if alt:
        cell.fill = PatternFill("solid", fgColor=CREMA)
    cell.border = Border(
        bottom=Side(style="hair", color="D8D2C4"),
        right=Side(style="hair", color="D8D2C4"),
    )


def hoja_maestra(wb, entradas):
    ws = wb.active
    ws.title = "Acervo (maestra)"

    # encabezados
    for col_idx, (_, header, width) in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        estilo_encabezado(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # filas
    for row_idx, e in enumerate(entradas, start=2):
        for col_idx, (key, _, _) in enumerate(COLUMNAS, start=1):
            v = e.get(key, "")
            if key == "etiquetas" and isinstance(v, list):
                v = " · ".join(v)
            if v is None:
                v = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            estilo_celda(cell, alt=(row_idx % 2 == 0))
            # destacar registros a verificar
            if key == "registro_digital" and "VERIFICAR" in str(v):
                cell.font = Font(name="Calibri", size=10, bold=True, color="8B1818")
            if key == "rubro":
                cell.font = Font(name="Calibri", size=10, bold=True)

    # congelar encabezado + 2 cols
    ws.freeze_panes = "C2"

    # autofiltro (no convertir a tabla porque chocan estilos con miles de filas)
    last_col = get_column_letter(len(COLUMNAS))
    ws.auto_filter.ref = f"A1:{last_col}{len(entradas)+1}"

    # altura mínima cómoda
    ws.row_dimensions[1].height = 30
    for r in range(2, len(entradas) + 2):
        ws.row_dimensions[r].height = 80

    return ws


def hoja_indice(wb, entradas):
    ws = wb.create_sheet("Índice", 0)

    md = ACERVO["metadata"]
    ws["A1"] = md["titulo"]
    ws["A1"].font = Font(name="Calibri", size=20, bold=True, color=ACENTO)
    ws["A2"] = md["subtitulo"]
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color=TINTA_SUAVE)
    ws["A4"] = f"Compilado: {md['fecha_compilacion']} · Versión {md['version']}"
    ws["A4"].font = Font(name="Calibri", size=10, color=TINTA_SUAVE)

    ws["A6"] = "ADVERTENCIA"
    ws["A6"].font = Font(name="Calibri", size=11, bold=True, color=ACENTO)
    ws["A7"] = md["advertencia"]
    ws["A7"].font = Font(name="Calibri", size=10, italic=True, color=TINTA_SUAVE)
    ws["A7"].alignment = Alignment(wrap_text=True, vertical="top")

    ws["A9"] = "JERARQUÍA LEGAL"
    ws["A9"].font = Font(name="Calibri", size=11, bold=True, color=ACENTO)
    ws["A10"] = md["jerarquia_legal"]
    ws["A10"].font = Font(name="Calibri", size=10, italic=True, color=TINTA_SUAVE)
    ws["A10"].alignment = Alignment(wrap_text=True, vertical="top")

    # estadística por materia y nivel
    ws["A13"] = "Distribución por materia"
    ws["A13"].font = Font(name="Calibri", size=12, bold=True, color=ACENTO)

    materias_cnt = {}
    for e in entradas:
        materias_cnt[e["materia_principal"]] = materias_cnt.get(e["materia_principal"], 0) + 1

    row = 15
    ws.cell(row=row, column=1, value="Materia").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Entradas").font = Font(bold=True)
    for m, c in sorted(materias_cnt.items(), key=lambda x: -x[1]):
        row += 1
        ws.cell(row=row, column=1, value=m).font = Font(size=10)
        ws.cell(row=row, column=2, value=c).font = Font(size=10)

    row += 2
    ws.cell(row=row, column=1, value="Distribución por jerarquía").font = Font(size=12, bold=True, color=ACENTO)
    row += 1
    jer_nombres = {
        1: "1 — Pleno SCJN",
        2: "2 — Salas SCJN",
        3: "3 — Plenos Regionales",
        4: "4 — Tribunales Colegiados",
        5: "5 — TSJ locales",
    }
    jer_cnt = {}
    for e in entradas:
        jer_cnt[e["jerarquia"]] = jer_cnt.get(e["jerarquia"], 0) + 1
    row += 1
    ws.cell(row=row, column=1, value="Nivel").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Entradas").font = Font(bold=True)
    for n in sorted(jer_cnt.keys()):
        row += 1
        ws.cell(row=row, column=1, value=jer_nombres.get(n, n)).font = Font(size=10)
        ws.cell(row=row, column=2, value=jer_cnt[n]).font = Font(size=10)

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 14

    return ws


def hojas_por_materia(wb, entradas):
    materias = sorted({e["materia_principal"] for e in entradas})
    for m in materias:
        sub = [e for e in entradas if e["materia_principal"] == m]
        nombre = m[:28].replace("/", "-")
        ws = wb.create_sheet(nombre)

        # Mismo formato que la maestra pero ya filtrado
        for col_idx, (_, header, width) in enumerate(COLUMNAS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            estilo_encabezado(cell)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, e in enumerate(sub, start=2):
            for col_idx, (key, _, _) in enumerate(COLUMNAS, start=1):
                v = e.get(key, "")
                if key == "etiquetas" and isinstance(v, list):
                    v = " · ".join(v)
                if v is None:
                    v = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=v)
                estilo_celda(cell, alt=(row_idx % 2 == 0))
                if key == "registro_digital" and "VERIFICAR" in str(v):
                    cell.font = Font(name="Calibri", size=10, bold=True, color="8B1818")
                if key == "rubro":
                    cell.font = Font(name="Calibri", size=10, bold=True)

        ws.freeze_panes = "C2"
        last_col = get_column_letter(len(COLUMNAS))
        ws.auto_filter.ref = f"A1:{last_col}{len(sub)+1}"
        ws.row_dimensions[1].height = 30
        for r in range(2, len(sub) + 2):
            ws.row_dimensions[r].height = 80


def make_xlsx():
    wb = Workbook()
    # eliminar hoja por defecto antes de crear las nuestras
    default = wb.active
    wb.remove(default)

    entradas = construir_entradas_normalizadas()
    entradas.sort(key=lambda x: (x["jerarquia"], x["materia_principal"], x["tema"], x["id"]))

    # 1. Índice
    ws_idx = wb.create_sheet("Índice")
    md = ACERVO["metadata"]
    ws_idx["A1"] = md["titulo"]
    ws_idx["A1"].font = Font(name="Calibri", size=20, bold=True, color=ACENTO)
    ws_idx["A2"] = md["subtitulo"]
    ws_idx["A2"].font = Font(name="Calibri", size=11, italic=True, color=TINTA_SUAVE)
    ws_idx["A4"] = f"Compilado: {md['fecha_compilacion']} · Versión {md['version']}"
    ws_idx["A4"].font = Font(name="Calibri", size=10, color=TINTA_SUAVE)
    ws_idx["A6"] = "ADVERTENCIA"
    ws_idx["A6"].font = Font(name="Calibri", size=11, bold=True, color=ACENTO)
    ws_idx["A7"] = md["advertencia"]
    ws_idx["A7"].font = Font(name="Calibri", size=10, italic=True, color=TINTA_SUAVE)
    ws_idx["A7"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_idx.row_dimensions[7].height = 60
    ws_idx["A9"] = "JERARQUÍA LEGAL"
    ws_idx["A9"].font = Font(name="Calibri", size=11, bold=True, color=ACENTO)
    ws_idx["A10"] = md["jerarquia_legal"]
    ws_idx["A10"].font = Font(name="Calibri", size=10, italic=True, color=TINTA_SUAVE)
    ws_idx["A10"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_idx.row_dimensions[10].height = 60

    materias_cnt = {}
    for e in entradas:
        materias_cnt[e["materia_principal"]] = materias_cnt.get(e["materia_principal"], 0) + 1
    ws_idx["A13"] = "Distribución por materia"
    ws_idx["A13"].font = Font(name="Calibri", size=12, bold=True, color=ACENTO)
    row = 15
    ws_idx.cell(row=row, column=1, value="Materia").font = Font(bold=True)
    ws_idx.cell(row=row, column=2, value="Entradas").font = Font(bold=True)
    for m, c in sorted(materias_cnt.items(), key=lambda x: -x[1]):
        row += 1
        ws_idx.cell(row=row, column=1, value=m).font = Font(size=10)
        ws_idx.cell(row=row, column=2, value=c).font = Font(size=10)

    row += 2
    ws_idx.cell(row=row, column=1, value="Distribución por jerarquía").font = Font(size=12, bold=True, color=ACENTO)
    row += 1
    jer_nombres = {
        1: "1 — Pleno SCJN",
        2: "2 — Salas SCJN",
        3: "3 — Plenos Regionales",
        4: "4 — Tribunales Colegiados",
        5: "5 — TSJ locales",
    }
    jer_cnt = {}
    for e in entradas:
        jer_cnt[e["jerarquia"]] = jer_cnt.get(e["jerarquia"], 0) + 1
    row += 1
    ws_idx.cell(row=row, column=1, value="Nivel").font = Font(bold=True)
    ws_idx.cell(row=row, column=2, value="Entradas").font = Font(bold=True)
    for n in sorted(jer_cnt.keys()):
        row += 1
        ws_idx.cell(row=row, column=1, value=jer_nombres.get(n, n)).font = Font(size=10)
        ws_idx.cell(row=row, column=2, value=jer_cnt[n]).font = Font(size=10)

    ws_idx.column_dimensions["A"].width = 60
    ws_idx.column_dimensions["B"].width = 14

    # 2. Hoja maestra
    ws_master = wb.create_sheet("Acervo (maestra)")
    for col_idx, (_, header, width) in enumerate(COLUMNAS, start=1):
        cell = ws_master.cell(row=1, column=col_idx, value=header)
        estilo_encabezado(cell)
        ws_master.column_dimensions[get_column_letter(col_idx)].width = width
    for row_idx, e in enumerate(entradas, start=2):
        for col_idx, (key, _, _) in enumerate(COLUMNAS, start=1):
            v = e.get(key, "")
            if key == "etiquetas" and isinstance(v, list):
                v = " · ".join(v)
            if v is None:
                v = ""
            cell = ws_master.cell(row=row_idx, column=col_idx, value=v)
            estilo_celda(cell, alt=(row_idx % 2 == 0))
            if key == "registro_digital" and "VERIFICAR" in str(v):
                cell.font = Font(name="Calibri", size=10, bold=True, color="8B1818")
            if key == "rubro":
                cell.font = Font(name="Calibri", size=10, bold=True)
    ws_master.freeze_panes = "C2"
    last_col = get_column_letter(len(COLUMNAS))
    ws_master.auto_filter.ref = f"A1:{last_col}{len(entradas)+1}"
    ws_master.row_dimensions[1].height = 30
    for r in range(2, len(entradas) + 2):
        ws_master.row_dimensions[r].height = 80

    # 3. Hojas por materia
    hojas_por_materia(wb, entradas)

    return wb


if __name__ == "__main__":
    out = Path("/sessions/brave-bold-goodall/mnt/outputs/acervo/Acervo_Jurisprudencial.xlsx")
    wb = make_xlsx()
    wb.save(out)
    print(f"XLSX generado: {out} ({out.stat().st_size:,} bytes)")
